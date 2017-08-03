import xml.etree.ElementTree as ET
import xlrd
import os
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename

def set_comlist():
    # global list to keep track of application b013's com list
    global b013_com_list # Needed to modify global copy of b013_comlist
    b013_com_list = []

def winpt_check(xcel_filename, directory, app, column, table_num, type):
    # xcel_filename - for the purpose of finding/opening the excel template file
    # directory - the path to get to the excel template file
    # app - the application currently in use
    # column - which column in the excel template you are wanting to check
    # table_num - which table in the SGConfig you are wanting to check
    # type - string value to indicate which type of point you are wanting to check

    # Try to read the file
    try:
        # Counters for WinPt status printing
        count = 0

        # Add the excel document to the passed-in directory
        filepath = directory + '/' + xcel_filename

        # Open the excel document for reading
        wbook = xlrd.open_workbook(filepath)

        # Read the specified excel sheet
        for sheet in wbook.sheet_names():
            if 'Sheet1' in sheet:
                wsheet_name = sheet

        wsheet = wbook.sheet_by_name(wsheet_name)

        # General Point Check
        print('\t\t', type, 'Points Check')

        # For all the records in the specified application table of the SGConfig
        for i, record in enumerate(app[table_num]):
            # Starting with the third point value in the specified excel column converted to a string for comparison reasons
            xl_value = str(wsheet.cell_value(i + 2, column))
            # The SGConfig point value
            check_value = record[0].get('Field_Value')
            # If the point is undefined in the SGConfig
            if record[0].get('Field_Value') == '(______) Undefined':
                print('\t\t\t', 'DNP Point', i, '<', type, '> Point is undefined.')
            # If the excel value is blank
            elif str(wsheet.cell_value(i+2, 1)) == '':
                print('\t\t\t', 'DNP Point', i, ': More SGConfig <', type, '> points than excel template <', type, '> points.')
                print('\t\t\t\t', 'Please match the number of excel points to the SGConfig.')
                break
            # If both the SGConfig and the excel have valid, string literal numbers
            else:
                # If the point in the SGConfig is less than three numbers
                if check_value[4] == '0':
                    # If the point in the SGConfig is one number
                    if check_value[5] == '0':
                        # If the first index of the excel value equals the one's index of the SGConfig
                        if xl_value[0] == (check_value[6]):
                            pass
                        # If the first index of the excel value does not equal the one's index of the SGConfig
                        else:
                            print('\t\t\t', 'DNP Point', i, '<',
                                  type, '> WinPt does not match the points list. Please refer to the SGConfig.')
                            count = count + 1  # Indicates that a WinPt does not match
                    # If the point in the SGConfig is two numbers
                    else:
                        # Check the first and second index of the excel value against the one's and ten's index of the SGConfig
                        if xl_value[0] + xl_value[1] == (check_value[5] + check_value[6]):
                            pass
                        # If the values do not match
                        else:
                            print('\t\t\t', 'DNP Point', i, '<',
                                  type, '> WinPt does not match the points list. Please refer to the SGConfig.')
                            count = count + 1  # Indicates that a WinPt does not match
                # If the point in the SGConfig is three numbers
                else:
                    # If they match, do nothing
                    if xl_value[0] + xl_value[1] + xl_value[2] == (check_value[4] + check_value[5] + check_value[6]):
                        pass
                    # If the values do not match
                    else:
                        print('\t\t\t', 'DNP Point', i, '<',
                                  type, '> WinPt does not match the points list. Please refer to the SGConfig.')
                        count = count + 1  # Indicates that a WinPt does not match

        # If all WinPts match, print statement
        if count == 0:
            print('\t\t\t', 'All <', type, '> WinPts match.')
        else:
            pass

    # PyCharm presents an error if the excel file is open. You have to close the document before running the program
    except Exception:
        print('\t\t\t', 'Error: Cannot read the file when it is open.')

def d20mpp_check(xml_filename, directory):
    tree = ET.parse(os.path.join(directory, xml_filename))
    root = tree.getroot()

    filename = askopenfilename(title='Select QC Doc to edit')

    ws_name = 'D20++ QC Doc'

    # For the purposes of editing the QC Doc
    print('\t', 'You have selected', filename, 'for editing.')
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name(ws_name)

    ws['B3'].value = xml_filename

    # Check all of these applications

    # Print the part number. For D20M++, the number should be 526-1006
    print(root[0][0][1][0].get('Part_Number'), '-', root[0][0].get('Device_Type') + 'M++')
    for app in root[0][0][1][0]:
        if app.get('Application_Identifier') == 'A003':
            a003_check(app, ws)
        if app.get('Application_Identifier') == 'A020':
            a020_check(app, ws)
        if app.get('Application_Identifier') == 'A026':
            a026_check(app, ws)
        if app.get('Application_Identifier') == 'A030':
            a030_check(app, ws)
        if app.get('Application_Identifier') == 'A083-0':
            a083_check(app, ws)
        if app.get('Application_Identifier') == 'B003':
           b003_check(app, ws)
        if app.get('Application_Identifier') == 'B013':
            b013_check(app, ws)
        if app.get('Application_Identifier') == 'B014':
            b014_check(app, ws)
        if app.get('Application_Identifier') == 'B015':
            b015_check(app, ws)
        if app.get('Application_Identifier') == 'B021':
            b021_check(app, ws)
        if app.get('Application_Identifier') == 'B023':
           b023_check(app, ws)

        ws['L' + str(198)].value = 'User Check'
        ws['L' + str(219)].value = 'User Check'
        ws['L' + str(222)].value = 'User Check'

        # Save Workbook Changes
        wb.save(filename)

def a003_check(app, ws):
    # Check SOE
    # Check Offline Condition
    # Check Contact BUR/BASE Time

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[6].get('Table_Identifier'), ':', app[6].get('Table_Name'), 'Table')
        # Tracking Variable
        count = 0
        # Loop through the specified table
        for i, record in enumerate(app[6]):
            if app[6][0][2].get('Field_Value') == 'No':
                pass  # Do nothing
            else:
                count = count + 1  # Increment for tracking purposes

        # Print Statement if an SOE Variable Differs
        if count == 0:
            print('\t\t', app[6][0][2].get('Field_Name'), ':', app[6][0][2].get('Field_Value'))
        else:
            print('An SOE value differs from the rest. Please check the SGConfig.')

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'), ':', app[1].get('Table_Name'), 'Table')
        # Loop through the specified table
        for i, record in enumerate(app[1]):
            print('\t\t', record[16].get('Field_Name'), ':', record[16].get('Field_Value'))

        print('\t', app[7].get('Table_Identifier'), ':', app[7].get('Table_Name'), 'Table')
        # Tracking variable
        count2 = 0
        # Loop through the specified table
        for i, record in enumerate(app[7]):
            if app[7][0][1].get('Field_Value') == '500':
                count2
            else:
                count2 = count2 + 1  # Increment for tracking purposes
        # Print Statement if a Contact Dur/Base Time variable differs
        if count2 == 0:
            print('\t\t', app[7][0][1].get('Field_Name'), ':', app[7][0][1].get('Field_Value'))
        else:
            print('A Contact Dur/Base Time value differs from the rest. Please check the SGConfig.')

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')
        # Edit excel document column "L" row "136"
        ws['L' + str(186)].value = "Application A003 is disabled for this site."
        ws['L' + str(187)].value = "Application A003 is disabled for this site."
        ws['L' + str(190)].value = "Application A003 is disabled for this site."

def a020_check(app, ws):
    # Check RE-INIT Interval

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))
        # Loop through the SRU table
        count = 0
        for i, record in enumerate(app[1]):
            print('\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))
            if record[4].get('Field_Value') == 60:
                pass
            else:
                count = count + 1

        if count == 0:
            ws['L' + str(40)].value = app[1][0][4].get('Field_Value')
        else:
            ws['L' + str(40)].value = 'One of the Re-init intervals does not match. Refer to the SGConfig.'

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')
        ws['L' + str(40)].value = "Application A020 is disabled for this site."

def a026_check(app, ws):
    # Check Point Type
    # Check System Point
    # Check Comm Event Point
    # Check Normal State


    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':

        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'), ':', app[0].get('Table_Name'), 'Table')
        # For the purpose of printing to the QC Doc
        row_number = 138
        ws['B' + str(row_number)].value = 'Record Point'
        ws['C' + str(row_number)].value = app[0][0][0].get('Field_Name')
        ws['D' + str(row_number)].value = app[0][0][1].get('Field_Name')
        ws['E' + str(row_number)].value = app[0][0][2].get('Field_Name')
        ws['F' + str(row_number)].value = app[0][0][3].get('Field_Name')

        # Loop through the Communication Events table
        for i, record in enumerate(app[0]):
            row_number = row_number + 1
            if row_number < 170:
                print('\t\t', i, ':')
                ws['B' + str(row_number)].value = i
                print('\t\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # Point Type
                ws['C' + str(row_number)].value = record[0].get('Field_Value')
                print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # System Point
                ws['D' + str(row_number)].value = record[1].get('Field_Value')
                print('\t\t\t', record[2].get('Field_Name'), ':', record[2].get('Field_Value'))  # Comm Even Point
                ws['E' + str(row_number)].value = record[2].get('Field_Value')
                print('\t\t\t', record[3].get('Field_Name'), ':', record[3].get('Field_Value'))  # Normal State
                ws['F' + str(row_number)].value = record[3].get('Field_Value')

            ws['L' + str(138)].value = "**May not contain all values. If you need them all, please refer to the SGConfig**"

        # Check SOE Enable
        # Check COS Enable

        # For the purpose of printing the SOE and COS to the QC Doc
        row_number = 171

        print('\t', app[2].get('Table_Identifier'), ':', app[2].get('Table_Name'), 'Table')
        # Loop through the DCA Configuration table
        for i, record in enumerate(app[2]):
            ws['L' + str(row_number)].value = record[1].get('Field_Value')
            ws['L' + str(row_number + 1)].value = record[2].get('Field_Value')
            print('\t\t\t', record[1].get('Field_Name'), ':', ws['L' + str(row_number)].value)
            print('\t\t\t', record[2].get('Field_Name'), ':', ws['L' + str(row_number + 1)].value)
            # print('\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # SOE Enable
            # print('\t\t', record[2].get('Field_Name'), ':', record[2].get('Field_Value'))  # COS Enable
            # The SOE and COS Enable values cannot both be "Yes"
            if record[1].get('Field_Value') == record[2].get('Field_Value'):
                print('\t\t', '** These values are not supposed to be the same. See the SGConfig. **')

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def a030_check(app, ws):
    # Check Time Sync Wait
    # Check Status/ACC Freeze
    # Check ACC Freeze/ Controls

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'), ':', app[1].get('Table_Name'), 'Table')

        # For the purpose of printing to the QC Doc
        test_value = app[1][0][1].get('Field_Value')  # The first record value
        count = 0

        # Loop through the DTA Misc Parameters table
        for i, record in enumerate(app[1]):
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))
            if record[1].get('Field_Value') == test_value:
                pass
            else:
                count = count + 1

        # For the purpose of printing to the QC Doc for Time Sync Wait, Status/ACC Freeze, and ACC Freeze/Controls
        row_number = 179

        if count == 0:
            ws['L' + str(row_number)].value = app[1][0][1].get('Field_Value')
        else:
            ws['L' + str(row_number)].value = 'A Time Sync Wait value differs. Please refer to the SGConfig.'

        # Print the table identifier followed by the table name for clarity
        print('\t', app[2].get('Table_Identifier'), ':', app[2].get('Table_Name'), 'Table')
        try:
            app[2][0].get('Record_Number')
            # Loop through the Status/ACC Freeze table
            for i, record in enumerate(app[2]):
                print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
                print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))
            ws['L' + str(row_number + 1)].value = 'There are STATUS/ACC FREEZE values. ' \
                                                      'Please remove them from the SGConfig.'
        except IndexError:
            print('\t\t', '<no entries>')
            ws['L' + str(row_number + 1)].value = 'There are no STATUS/ACC FREEZE values.'

        # Print the table identifier followed by the table name for clarity
        print('\t', app[3].get('Table_Identifier'), ':', app[3].get('Table_Name'), 'Table')
        try:
            app[3][0].get('Record_Number')
            # Loop through the ACC Freeze/Controls table
            for i, record in enumerate(app[3]):
                print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
                print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))
            ws['L' + str(row_number + 2)].value = 'There are ACC FREEZE/CONTROLS values. ' \
                                                  'Please remove them from the SGConfig.'
        # If there are no records
        except IndexError:
            print('\t\t', '<no entries>')
            ws['L' + str(row_number + 2)].value = 'There are no ACC FREEZE/CONTROLS values.'

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def a083_check(app, ws):
    # Check That All Points Have Event Types = Both

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        print('\t\t D20 Calculator: Digital Inputs table does not have event types')
        ws['L42'].value = 'D20 Calculator: Digital Inputs table does not have event types.'
        # for record in app[2][0]:
        #     print('\t Calc', record.get('Record_Number'))
        #     print('\t Calc', record.get('Record_Number'), '-', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b003_check(app, ws):
    # The XML export does not contain the report deadband.

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))
        print('\t', 'Report Deadband not in XML')
        return

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b013_check(app, ws):
    # Check Baud Rate
    # Check DCD, RTS, & CTS
    # Check DCD to RX Enable Time
    # Check RTS Preamble
    # Check RTS Postamble
    # Check Max Frame Size
    # Check Transmit Retries
    # Check Transmit Buffers
    # Check Receive Buffers
    # Check Confirm Timeout
    # Check Response Timeout

    # For the purpose of printing to the QC Doc
    row_num = 206

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'), ':', app[0].get('Table_Name'), 'Table')
        # Port Com Global Variable, enable editing
        set_comlist()

        # Indication Counts
        dcdcount = 0
        rxencount = 0
        rtsprecount = 0
        rtspostcount = 0
        maxframecount = 0
        trretriecount = 0
        trbuffcount = 0
        recbuffcount = 0
        contimcount = 0
        restimcount = 0
        # Reference Values
        dcdval = app[0][0][3].get('Field_Value')
        rxenval = app[0][0][6].get('Field_Value')
        rtspreval = app[0][0][7].get('Field_Value')
        rtspostval = app[0][0][8].get('Field_Value')
        maxframeval = app[0][0][9].get('Field_Value')
        trretrieval = app[0][0][10].get('Field_Value')
        trbuffval = app[0][0][11].get('Field_Value')
        recbuffval = app[0][0][12].get('Field_Value')
        contimval = app[0][0][13].get('Field_Value')
        restimval = app[0][0][14].get('Field_Value')

        # Loop through the Port Configuration table
        for i, record in enumerate(app[0]):
            print('\t\t', i, ':')
            print('\t\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # Port
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # Baud Rate
            print('\t\t\t', record[3].get('Field_Name'), ':', record[3].get('Field_Value'))  # DCD
            print('\t\t\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))  # RTS
            print('\t\t\t', record[5].get('Field_Name'), ':', record[5].get('Field_Value'))  # CTS
            if record[3].get('Field_Value') != dcdval:
                dcdcount = dcdcount + 1
            elif record[4].get('Field_Value') != dcdval:
                dcdcount = dcdcount + 1
            elif record[5].get('Field_Value') != dcdval:
                dcdcount = dcdcount + 1
            print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))  # DCD to Rx Enable Time
            if record[6].get('Field_Value') != rxenval:
                rxencount = rxencount + 1
            print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))  # RTS Preamble
            if record[7].get('Field_Value') != rtspreval:
                rtsprecount = rtsprecount + 1
            print('\t\t\t', record[8].get('Field_Name'), ':', record[8].get('Field_Value'))  # RTS Postamble
            if record[8].get('Field_Value') != rtspostval:
                rtspostcount = rtspostcount + 1
            print('\t\t\t', record[9].get('Field_Name'), ':', record[9].get('Field_Value'))  # Max Frame Size
            if record[9].get('Field_Value') != maxframeval:
                maxframecount = maxframecount + 1
            print('\t\t\t', record[10].get('Field_Name'), ':', record[10].get('Field_Value'))  # Transmit Retries
            if record[10].get('Field_Value') != trretrieval:
                trretriecount = trretriecount + 1
            print('\t\t\t', record[11].get('Field_Name'), ':', record[11].get('Field_Value'))  # Transmit Buffers
            if record[11].get('Field_Value') != trbuffval:
                trbuffcount = trbuffcount + 1
            print('\t\t\t', record[12].get('Field_Name'), ':', record[12].get('Field_Value'))  # Receive Buffers
            if record[12].get('Field_Value') != recbuffval:
                recbuffcount = recbuffcount + 1
            print('\t\t\t', record[13].get('Field_Name'), ':', record[13].get('Field_Value'))  # Confirm Timeout
            if record[13].get('Field_Value') != contimval:
                contimcount = contimcount + 1
            print('\t\t\t', record[14].get('Field_Name'), ':', record[14].get('Field_Value'))  # Response Timeout
            if record[14].get('Field_Value') != restimval:
                restimcount = restimcount + 1

            # Append Port Com to list
            b013_com_list.append((record[0].get('Field_Value')))

        # QC Doc DCD, RTS, CTS
        if dcdcount == 0:
            ws['L' + str(row_num)].value = dcdval
        else:
            ws['L' + str(row_num)].value = 'A DCD, RTS, or CTS value differs. Please refer to the SGConfig.'
        # QC Doc Baud Rate
        ws['L' + str(176)].value = 'Please refer to the SGConfig for the Baud Rate.'
        # QC Doc Rx Enable Time
        if rxencount == 0:
            ws['L' + str(row_num + 1)].value = rxenval
        else:
            ws['L' + str(row_num + 1)].value = 'A DCD to Rx Enable Time value differs. Please refer to the SGConfig.'
        # QC Doc RTS Preamble
        if rtsprecount == 0:
            ws['L' + str(row_num + 2)].value = rtspreval
        else:
            ws['L' + str(row_num + 2)].value = 'A RTS Preamble value differs. Please refer to the SGConfig.'
        # QC Doc RTS Postamble
        if rtspostcount == 0:
            ws['L' + str(row_num + 3)].value = rtspostval
        else:
            ws['L' + str(row_num + 3)].value = 'A RTS Postamble value differs. Please refer to the SGConfig.'
        # QC Doc Max Frame Size
        if maxframecount == 0:
            ws['L' + str(row_num + 4)].value = maxframeval
        else:
            ws['L' + str(row_num + 4)].value = 'A Max Frame Size value differs. Please refer to the SGConfig.'
        # QC Doc Transmit Retries
        if trretriecount == 0:
            ws['L' + str(row_num + 5)].value = trretrieval
        else:
            ws['L' + str(row_num + 5)].value = 'A Transmit Retries value differs. Please refer to the SGConfig.'
        # QC Doc Transmit Buffers
        if trbuffcount == 0:
            ws['L' + str(row_num + 6)].value = trbuffval
        else:
            ws['L' + str(row_num + 6)].value = 'A Transmit Buffers value differs. Please refer to the SGConfig.'
        # QC Doc Receive Buffers
        if recbuffcount == 0:
            ws['L' + str(row_num + 7)].value = recbuffval
        else:
            ws['L' + str(row_num + 7)].value = 'A Receive Buffers value differs. Please refer to the SGConfig.'
        # QC Doc Confirm Timeout
        if contimcount == 0:
            ws['L' + str(row_num + 8)].value = contimval
        else:
            ws['L' + str(row_num + 8)].value = 'A Confirm Timeout value differs. Please refer to the SGConfig.'
        # QC Doc Response Timeout
        if restimcount == 0:
            ws['L' + str(row_num + 9)].value = restimval
        else:
            ws['L' + str(row_num + 9)].value = 'A Response Timeout value differs. Please refer to the SGConfig.'

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')
        ws['L' + str(row_num)].value = 'Application B013 is disabled (Rows 206 - 215).'

def b014_check(app, ws):
    # Check SOE BUFFER SIZE
    # Check SOE LOCATION

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'), ':', app[1].get('Table_Name'))

        # For the purpose of printing to the QC Doc
        row_num = 67
        sbuffer_count = 0
        slocation_count = 0
        sbuffer_val = app[1][0][0].get('Field_Value')
        slocation_val = app[1][0][4][0][0][0].get('Field_Value')

        # Loop through the Buffer Configuration table
        for i, record in enumerate(app[1]):
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # SOE Buffer Size
            if record[0].get('Field_Value') == sbuffer_val:
                pass
            else:
                sbuffer_count = sbuffer_count + 1
            print('\t\t', record[4][0][0][0].get('Field_Name'), ':', record[4][0][0][0].get('Field_Value'))  # SOE Location
            if record[4][0][0][0].get('Field_Value') == slocation_val:
                pass
            else:
                slocation_count = slocation_count + 1
        # Print the SOE BUFFER SIZE to the QC Doc
        if sbuffer_count == 0:
            ws['L' + str(row_num)].value = sbuffer_val
        else:
            ws['L' + str(row_num)].value = 'Not all SOE BUFFER SIZE values are the same. Please refer to the SGConfig.'
        # Print the SOE LOCATION to the QC Doc
        if slocation_count == 0:
            ws['L' + str(row_num + 1)].value = slocation_val
        else:
            ws['L' + str(row_num + 1)].value = 'Not all SOE LOCATION values are the same. Please refer to the SGConfig.'

        # Check the Standard UTC Offset
        # Check the DST Offset

        # Print the table identifier followed by the table name for clarity
        print('\t', app[4].get('Table_Identifier'), ':', app[4].get('Table_Name'), 'Table')

        # For the purpose of printing to the QC Doc
        utc_count = 0
        dst_count = 0
        utc_val = app[4][0][11].get('Field_Value')
        dst_val = app[4][0][12].get('Field_Value')

        # Loop through the Daylight Savings Time table
        for i, record in enumerate(app[4]):
            print('\t\t', record[11].get('Field_Name'), ':', record[11].get('Field_Value'))  # Standard UTC Offset
            if record[11].get('Field_Value') != utc_val:
                utc_count = utc_count + 1
            print('\t\t', record[12].get('Field_Name'), ':', record[12].get('Field_Value'))  # DST Offset
            if record[12].get('Field_Value') != dst_val:
                dst_count = dst_count + 1

        if utc_count == 0:
            ws['L' + str(row_num + 2)].value = utc_val
        else:
            ws['L' + str(row_num + 2)].value = 'Not all Standard UTC Offset values are the same. ' \
                                               'Please refer to the SGConfig.'
        if dst_count == 0:
            ws['L' + str(row_num + 3)].value = dst_val
        else:
            ws['L' + str(row_num + 3)].value = 'Not all DST Offset values are the same. Please refer to the SGConfig.'

        # Check User Name = something
        # Check Password = something
        # Check Control Password = something

        # Print the table identifier followed by the table name for clarity
        print('\t', app[2].get('Table_Identifier'), ':', app[2].get('Table_Name'), 'Table')
        # Loop through the User Configuration table
        for i, record in enumerate(app[2]):
            record_num = int(record.get('Record_Number')) - 1
            print('\t\t', 'Record ', record_num, ':')
            print('\t\t\t', record[5].get('Field_Name'), ':', record[5].get('Field_Value'))  # User Name
            print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))  # Password
            print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))  # Control Password

        # For the purpose of printing to the QC Doc
        ws['L71'].value = "Please refer to the SGConfig."
        row_num = 78

        # Check Welcome Message

        # Print the table identifier followed by the table name for clarity
        print('\t', app[5].get('Table_Identifier'), ':', app[5].get('Table_Name'), 'Table')
        # Loop through the Welcome Message table
        for i, record in enumerate(app[5]):
            print('\t\t', record[0].get('Field_Name'), record[0].get('Field_Value'), record[3].get('Field_Name'),
                  ':', record[3].get('Field_Value'))
            ws['C' + str(row_num)].value = record[0].get('Field_Value')
            ws['D' + str(row_num)].value = record[3].get('Field_Value')
            row_num = row_num + 1



    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b015_check(app, ws):
    # Check Bridgeman app

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Count the number of remote DNP devices
        num_dnp_dev = 0
        for record in app[2][0]:
            num_dnp_dev += 1
        print('\t', num_dnp_dev, 'remote DNP devices')

        # For the purpose of printing to the QC Doc
        rxcount = 0
        rxval = app[0][0][1].get('Field_Value')

        for record in app[0]:
            print('\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # Number of Rx Buffers
            if record[1].get('Field_Value') != rxval:
                rxcount = rxcount + 1

        if rxcount == 0:
            ws['L' + str(110)].value = rxval
        else:
            ws['L' + str(110)].value = 'Not all Number of RX Buffers values are the same. Please refer to the SGConfig.'

        # For the purpose of printing to the Local Application table of the QC Doc
        row_number = 114

        print('\t', 'Local Application Table [LAN Address(Hex), Data Link Channel]')
        # Loop through the Local Application table
        for i, record in enumerate(app[2]):
            print('\t\t', record[0].get('Field_Value'), '(x', record[3].get('Field_Value'), ')',
                  record[2].get('Field_Value'), ':', b013_com_list[i])
            ws['D' + str(row_number)].value = record[0].get('Field_Value')
            ws['E' + str(row_number)].value = record[2].get('Field_Value')
            ws['F' + str(row_number)].value = record[3].get('Field_Value')
            ws['L' + str(row_number)].value = b013_com_list[i]
            row_number = row_number + 1

        # Indication of COM number
        ws['L' + str(113)].value = 'Corresponding COM Port'
        ws['L' + str(123)].value = 'Corresponding COM Port'

        # For the purpose of printing to the Remote Application Table of the QC Doc
        row_number = 124
        txcount = 0
        txval = app[3][0][4].get('Field_Value')

        print('\t', 'Remote Application Table [LAN Address(Hex), Data Link Channel]')
        # Check TXT Delay to Appl.
        # Loop through the Remote Application table
        for i, record in enumerate(app[3]):
            print('\t\t', record[0].get('Field_Value'), '(x', record[3].get('Field_Value'), ')',
                  record[2].get('Field_Value'), '   ', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))
            ws['D' + str(row_number)].value = record[0].get('Field_Value')
            ws['E' + str(row_number)].value = record[2].get('Field_Value')
            ws['F' + str(row_number)].value = record[3].get('Field_Value')
            ws['L' + str(row_number)].value = b013_com_list[i]
            if record[4].get('Field_Value') != txval:
                txcount = txcount + 1
            row_number = row_number + 1

        if txcount == 0:
            ws['L' + str(205)].value = txval
        else:
            ws['L' + str(205)].value = 'Refer to the SGConfig for TX DELAY TO APPL. values.'

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b021_check(app, ws):
    # Check Datalink Confirm
    # Check Idle Report Period

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'), ':', app[0].get('Table_Name'), 'Table')

        # For the purpose of printing to the QC Doc
        datalink_control_value = app[0][0][14].get('Field_Value')
        idlereport_control_value = app[0][0][11][0][0][5].get('Field_Value')
        datalink_count = 0
        idlereport_count = 0
        row_number = 183

        # Loop through the DPA Configuration table
        for i, record in enumerate(app[0]):
            print('\t\t', i, ':')
            print('\t\t\t', record[14].get('Field_Name'), ':', record[14].get('Field_Value'))  # Data Link Confirm
            if record[14].get('Field_Value') == datalink_control_value:
                pass
            else:
                datalink_count = datalink_count + 1
            print('\t\t\t', record[11].get('Field_Name'))  # Unsolicited Response
            print('\t\t\t\t', record[11][0][0][5].get('Field_Name'), ':', record[11][0][0][5].get('Field_Value'))
                                                                                               # Idle Report Period
            if record[11][0][0][5].get('Field_Value') == idlereport_control_value:
                pass
            else:
                idlereport_count = idlereport_count + 1

        if datalink_count == 0:
            ws['L' + str(row_number)].value = datalink_control_value
        else:
            ws['L' + str(row_number)].value = 'There is a DATALINK CONFIRM value that is not disabled. ' \
                                              'Please refer to the SGConfig.'
        if idlereport_count == 0:
            ws['L' + str(row_number + 1)].value = idlereport_control_value
        else:
            ws['L' + str(row_number + 1)].value = 'There is an Idle Report Period value that differs. ' \
                                              'Please refer to the SGConfig.'

        # Compare the Winpoints from the points list to what's programmed in the D20

        # Show an "Open" dialog box and return the path to the selected file
        filename = askopenfilename(title='Select EXCEL D20 DNP Map WinPt Check')

        # Put in the path to the excel template file
        directory = os.path.expanduser(
            os.path.join('~', 'Documents', 'GitHub', 'FE-D20_Checker', 'Example D20 XML', 'D20MEII'))

        # Determine that the XCEL template's filename is D20 DNP Map WinPt Check
        for thing in os.listdir(directory):
            # Excel template should be named D20 DNP Map WinPt Check
            if 'D20 DNP Map WinPt Check' in thing:
                xcel_filename = thing
                print('\t', xcel_filename)

            # Try to read the file
            try:
                # Add the filename to the directory
                filepath = directory + '/' + xcel_filename

                # Open the excel document for reading
                wbook = xlrd.open_workbook(filepath)

                # Read the specified excel sheet
                for sheet in wbook.sheet_names():
                    if 'Sheet1' in sheet:
                        wsheet_name = sheet

                wsheet = wbook.sheet_by_name(wsheet_name)

                # Determine which column the specified points are in
                for i, cell in enumerate(wsheet.row(1)):
                    if cell.value == 'DNP INDEX':
                        dnp_index = i
                    elif cell.value == 'STATUS':
                        status_index = i
                    elif cell.value == 'ANALOG':
                        analog_index = i
                    elif cell.value == 'CONTROL':
                        control_index = i

            # PyCharm presents an error if the excel file is open. You have to close the document
            # before running the program
            except Exception:
                print('\t\t\t', 'Error: Cannot read the file.')

        # Call the WinPt check function for Status, Analog, and Control points respectively
        winpt_check(xcel_filename, directory, app, status_index, 3, 'Status')
        winpt_check(xcel_filename, directory, app, analog_index, 6, 'Analog')
        winpt_check(xcel_filename, directory, app, control_index, 4, 'Control')

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b023_check(app, ws):
    # <app>
    #   <table "B023_PNT">
    #   There is no B023_POL
    #   <table "B023_DEV">
    #   <table "B023_CFG">


    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[2].get('Table_Identifier'))
        b023_pnt_list = []

        # For the use of putting the following values into the QC Doc
        pnt_location_number = 23
        cfg_row_number = 33
        dev_row_number = 216

        # Loop through the Device Point Map table
        for i, record in enumerate(app[2]):

            ws['E' + str(pnt_location_number)].value = record[0].get('Field_Value')
            ws['F' + str(pnt_location_number)].value = record[1].get('Field_Value')
            print('\t\t', i, '-', record[0].get('Field_Value'), ':', record[1].get('Field_Value'))
            b023_pnt_list.append((record[0].get('Field_Value'), record[1].get('Field_Value')))
            pnt_location_number = pnt_location_number + 1

        print('\t', 'B023_POL')
        #b023_pol_list = []
        print('\t\t D20 DNP DCA does not have a POL list')
        # for i, record in enumerate(app[3]):
        #     print('\t\t', 'Record', ':', i)
        #     print('\t\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
        #     print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))
        #     print('\t\t\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))
        #     print('\t\t\t', record[5].get('Field_Name'), ':', record[5].get('Field_Value'))
        #     print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))
        #     print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))
        #     print('\t\t\t', record[8].get('Field_Name'), ':', record[8].get('Field_Value'))
        #     b023_pol_list.append((i, record[0].get('Field_Value')))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'))
        b023_dev_list = []

        # For the purpose of printing to the QC Doc
        pollCount = 0
        integCount = 0
        offCount = 0
        failCount = 0
        timeCount = 0
        dataCount = 0

        pollVal = app[1][0][5][0][0][3].get('Field_Value')
        integVal = app[1][0][6].get('Field_Value')
        offVal = app[1][0][3][0][0][8].get('Field_Value')
        failVal = app[1][0][7].get('Field_Value')
        timeVal = app[1][0][3][0][0][4].get('Field_Value')
        dataVal = app[1][0][3][0][0][7].get('Field_Value')

        # Loop through the Device Configuration table
        for i, record in enumerate(app[1]):
            # Application Address
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
            # Poll Interval (s)
            print('\t\t\t', record[5][0][0][3].get('Field_Name'), ':', record[5][0][0][3].get('Field_Value'))
            if record[5][0][0][3].get('Field_Value') != pollVal:
                pollCount = pollCount + 1
            # Integrity Poll Interval
            print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))
            if record[6].get('Field_Value') != integVal:
                integCount = integCount + 1
            # Offline After Fail
            print('\t\t\t', record[3][0][0][8].get('Field_Name'), ':', record[3][0][0][8].get('Field_Value'))
            if record[3][0][0][8].get('Field_Value') != offVal:
                offCount = offCount + 1
            # Failures For Bad Channel
            print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))
            if record[7].get('Field_Value') != failVal:
                failCount = failCount + 1
            # Time Syncing
            print('\t\t\t', record[3][0][0][4].get('Field_Name'), ':', record[3][0][0][4].get('Field_Value'))
            if record[3][0][0][4].get('Field_Value') != timeVal:
                timeCount = timeCount + 1
            # Data Link CFM Required
            print('\t\t\t', record[3][0][0][7].get('Field_Name'), ':', record[3][0][0][7].get('Field_Value'))
            if record[3][0][0][7].get('Field_Value') != dataVal:
                dataCount = dataCount + 1
            # First Point Record
            print('\t\t\t', record[9][0][0][2].get('Field_Name'), ':', record[9][0][0][2].get('Field_Value'))
            # Number of Point Records
            print('\t\t\t', record[9][0][0][3].get('Field_Name'), ':', record[9][0][0][3].get('Field_Value'))
            for index in range(int(record[9][0][0][2].get('Field_Value')),
                               int(record[9][0][0][2].get('Field_Value')) + int(
                                       record[9][0][0][3].get('Field_Value'))):
                print('\t\t\t\t', b023_pnt_list[index])
            # There are no POLL Records for the D20 DNP DCA
            # print('\t\t\t', record[8][0][0][4].get('Field_Name'), ':', record[8][0][0][4].get('Field_Value'))
            # print('\t\t\t', record[8][0][0][5].get('Field_Name'), ':', record[8][0][0][5].get('Field_Value'))
            # for index in range(int(record[9][0][0][4].get('Field_Value')),
            #                    int(record[9][0][0][4].get('Field_Value')) + int(
            #                            record[9][0][0][5].get('Field_Value'))):
            #     print('\t\t\t\t', b023_pol_list[index])
            # print('\t\t\t', record[9][0][0][5].get('Field_Name'), ':', record[9][0][0][5].get('Field_Value'))
            b023_dev_list.append(record[0].get('Field_Value'))

        if pollCount == 0:
            ws['L' + str(cfg_row_number + 1)].value = pollVal
        else:
            ws['L' + str(cfg_row_number + 1)].value = 'Some of the POLL INTERVAL values differ. ' \
                                                      'Please refer to the SGConfig.'
        if integCount == 0:
            ws['L' + str(cfg_row_number + 2)].value = integVal
        else:
            ws['L' + str(cfg_row_number + 2)].value = 'Some of the INTEGRITY POLL INTERVAL values differ. ' \
                                                      'Please refer to the SGConfig.'
        if offCount == 0:
            ws['L' + str(cfg_row_number + 3)].value = offVal
        else:
            ws['L' + str(cfg_row_number + 3)].value = 'Some of the OFFLINE AFTER FAIL values differ. ' \
                                                      'Please refer to the SGConfig.'
        if failCount == 0:
            ws['L' + str(cfg_row_number + 4)].value = failVal
        else:
            ws['L' + str(cfg_row_number + 4)].value = 'Some of the FAILURES FOR BAD CHAN. values differ. ' \
                                                      'Please refer to the SGConfig.'
        if timeCount == 0:
            ws['L' + str(cfg_row_number + 5)].value = timeVal
        else:
            ws['L' + str(cfg_row_number + 5)].value = 'Some of the TIMESYNCING values differ.' \
                                                      ' Please refer to the SGConfig.'
        if dataCount == 0:
            ws['L' + str(dev_row_number)].value = dataVal
        else:
            ws['L' + str(dev_row_number)].value = 'Some of the Data Link CFM Required' \
                                                  ' values differ. Please refer to the SGConfig.'

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'))

        # For the purpose of printing to the QC Doc
        restart_count = 0
        restart_val = app[0][0][2].get('Field_Value')

        # Loop through the DCA Configuration table
        for i, record in enumerate(app[0]):
            #DCA Address
            print('\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))
            #Restart Delay
            print('\t\t\t', record[2].get('Field_Name'), ':', record[2].get('Field_Value'))
            if restart_val != record[2].get('Field_Value'):
                restart_count = restart_count + 1
            print('\t\t\t', 'Devices in DCA:')
            for index in range(int(record[8].get('Field_Value')),
                               int(record[8].get('Field_Value')) + int(record[9].get('Field_Value'))):
                print('\t\t\t\t', b023_dev_list[index])

        if restart_count == 0:
            ws['L' + str(cfg_row_number)].value = restart_val
        else:
            ws['L' + str(cfg_row_number)].value = 'One of the restart values differ. Please refer to the SGConfig.'
        return

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

