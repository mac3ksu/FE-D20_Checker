import xml.etree.ElementTree as ET
import xlrd
import os
from openpyxl import Workbook
from openpyxl import load_workbook


def set_comlist():
    # global list to keep track of application b013's com list
    global b013_com_list # Needed to modify global copy of globvar
    b013_com_list = []

def winpt_check(xcel_filename, directory, app, column, table_num, type):
    # xcel_filename - for the purpose of finding/opening the excel template file
    # directory - the path to get to the excel template file
    # app - the application currently in use
    # column - which column in the excel template you are wanting to check
    # table_num - which table in the SGConfig you are wanting to check
    # type - string value to indicate which type of point you are wanting to check

    # Try to read the file
    try:
        # Counters for WinPt status printing
        count = 0

        # Add the excel document to the passed-in directory
        filepath = directory + '/' + xcel_filename

        # Open the excel document for reading
        wbook = xlrd.open_workbook(filepath)
        #wbook = Workbook(filepath) # Just makes a new xcel doc?
        #wbook = load_workbook(filepath)  # Takes in a workbook

        # Read the specified excel sheet
        for sheet in wbook.sheet_names():
            if 'Sheet1' in sheet:
                wsheet_name = sheet

        wsheet = wbook.sheet_by_name(wsheet_name)

        # General Point Check
        print('\t\t', type, 'Points Check')

        # For all the records in the specified application table of the SGConfig
        for i, record in enumerate(app[table_num]):
            # Starting with the third point value in the specified excel column converted to a string for comparison reasons
            xl_value = str(wsheet.cell_value(i + 2, column))
            # The SGConfig point value
            check_value = record[0].get('Field_Value')
            # If the point is undefined in the SGConfig
            if record[0].get('Field_Value') == '(______) Undefined':
                print('\t\t\t', 'DNP Point', i, '<', type, '> Point is undefined.')
            # If the excel value is blank
            elif str(wsheet.cell_value(i+2, 1)) == '':
                print('\t\t\t', 'DNP Point', i, ': More SGConfig <', type, '> points than excel template <', type, '> points.')
                print('\t\t\t\t', 'Please match the number of excel points to the SGConfig.')
                break
            # If both the SGConfig and the excel have valid, string literal numbers
            else:
                # If the point in the SGConfig is less than three numbers
                if check_value[4] == '0':
                    # If the point in the SGConfig is one number
                    if check_value[5] == '0':
                        # If the first index of the excel value equals the one's index of the SGConfig
                        if xl_value[0] == (check_value[6]):
                            pass
                        # If the first index of the excel value does not equal the one's index of the SGConfig
                        else:
                            print('\t\t\t', 'DNP Point', i, '<',
                                  type, '> WinPt does not match the points list. Please refer to the SGConfig.')
                            count = count + 1  # Indicates that a WinPt does not match
                    # If the point in the SGConfig is two numbers
                    else:
                        # Check the first and second index of the excel value against the one's and ten's index of the SGConfig
                        if xl_value[0] + xl_value[1] == (check_value[5] + check_value[6]):
                            pass
                        # If the values do not match
                        else:
                            print('\t\t\t', 'DNP Point', i, '<',
                                  type, '> WinPt does not match the points list. Please refer to the SGConfig.')
                            count = count + 1  # Indicates that a WinPt does not match
                # If the point in the SGConfig is three numbers
                else:
                    # If they match, do nothing
                    if xl_value[0] + xl_value[1] + xl_value[2] == (check_value[4] + check_value[5] + check_value[6]):
                        pass
                    # If the values do not match
                    else:
                        print('\t\t\t', 'DNP Point', i, '<',
                                  type, '> WinPt does not match the points list. Please refer to the SGConfig.')
                        count = count + 1  # Indicates that a WinPt does not match

        # If all WinPts match, print statement
        if count == 0:
            print('\t\t\t', 'All <', type, '> WinPts match.')
        else:
            pass

    # PyCharm presents an error if the excel file is open. You have to close the document before running the program
    except Exception:
        print('\t\t\t', 'Error: Cannot read the file.')

def d20meII_check(xml_filename, directory):
    tree = ET.parse(os.path.join(directory, xml_filename))
    root = tree.getroot()

    # Check all of these applications

    # Print the part number. For D20MEII, the number should be 526-2007
    print(root[0][0][1][0].get('Part_Number'), '-', root[0][0].get('Device_Type') + 'MEII')
    for app in root[0][0][1][0]:
        if app.get('Application_Identifier') == 'A003':
            a003_check(app)
        if app.get('Application_Identifier') == 'A020':
            a020_check(app)
        if app.get('Application_Identifier') == 'A026-1':
            a026_check(app)
        if app.get('Application_Identifier') == 'A030':
            a030_check(app)
        if app.get('Application_Identifier') == 'A083-0':
            a083_check(app)
        if app.get('Application_Identifier') == 'B003':
            b003_check(app)
        if app.get('Application_Identifier') == 'B013':
            b013_check(app)
        if app.get('Application_Identifier') == 'B014-1':
            b014_check(app)
        if app.get('Application_Identifier') == 'B015':
            b015_check(app)
        if app.get('Application_Identifier') == 'B021':
            b021_check(app)
        if app.get('Application_Identifier') == 'B023':
            b023_check(app)

# Application A003 is NOT in the D20MEII XML File
def a003_check(app):
    # Check SOE
    # Check Offline Condition
    # Check Contact BUR/BASE Time

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[6].get('Table_Identifier'), ':', app[6].get('Table_Name'), 'Table')
        # Tracking Variable
        count = 0
        # Loop through the specified table
        for i, record in enumerate(app[6]):
            if app[6][0][2].get('Field_Value') == 'No':
                pass  # Do nothing
            else:
                count = count + 1  # Increment for tracking purposes

        # Print Statement if an SOE Variable Differs
        if count == 0:
            print('\t\t', app[6][0][2].get('Field_Name'), ':', app[6][0][2].get('Field_Value'))
        else:
            print('An SOE value differs from the rest. Please check the SGConfig.')

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'), ':', app[1].get('Table_Name'), 'Table')
        # Loop through the specified table
        for i, record in enumerate(app[1]):
            print('\t\t', record[16].get('Field_Name'), ':', record[16].get('Field_Value'))

        print('\t', app[7].get('Table_Identifier'), ':', app[7].get('Table_Name'), 'Table')
        # Tracking variable
        count2 = 0
        # Loop through the specified table
        for i, record in enumerate(app[7]):
            if app[7][0][1].get('Field_Value') == '500':
                count2
            else:
                count2 = count2 + 1  # Increment for tracking purposes
        # Print Statement if a Contact Dur/Base Time variable differs
        if count2 == 0:
            print('\t\t', app[7][0][1].get('Field_Name'), ':', app[7][0][1].get('Field_Value'))
        else:
            print('A Contact Dur/Base Time value differs from the rest. Please check the SGConfig.')

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def a020_check(app):
    # Check RE-INIT Interval

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))
        # Loop through the SRU table
        for i, record in enumerate(app[1]):
            print('\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def a026_check(app):
    # Check Operating Condition
    # Check Channel Type/Specifier
    # Check Status Point
    # Check Normal State
    # Check Start Point

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'), ':', app[0].get('Table_Name'), 'Table')
        # Loop through the Communication Events table
        for i, record in enumerate(app[0]):
            print('\t\t', i, ':')
            print('\t\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # Operating Condition
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # Channel/Type Specifier
            print('\t\t\t', record[2].get('Field_Name'), ':', record[2].get('Field_Value'))  # Status Point
            print('\t\t\t', record[3].get('Field_Name'), ':', record[3].get('Field_Value'))  # Normal State
            print('\t\t\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))  # Start Point

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def a030_check(app):
    # Check Time Sync Wait
    # Check Status/ACC Freeze
    # Check ACC Freeze/ Controls

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'), ':', app[1].get('Table_Name'), 'Table')
        # Loop through the DTA Misc Parameters table
        for i, record in enumerate(app[1]):
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[2].get('Table_Identifier'), ':', app[2].get('Table_Name'), 'Table')
        # Loop through the Status/ACC Freeze table
        for i, record in enumerate(app[2]):
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[3].get('Table_Identifier'), ':', app[3].get('Table_Name'), 'Table')
        # Loop through the ACC Freeze/Controls table
        for i, record in enumerate(app[3]):
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def a083_check(app):
    # Check all calc points have Event Types = Both

    # Check if Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[5].get('Table_Identifier'), ':', app[5].get('Table_Name'), 'Table')
        # Loop through the Digital Inputs table
        for record in app[5]:
            print('\t\t Calc', record.get('Record_Number'), '-', record[2].get('Field_Name'), ':',
                  record[2].get('Field_Value'))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b003_check(app):
    # Check if Application is Enabled
    if app.get('Enabled') == 'True':
        # The XML export does not contain the report deadband.
        print('B003 - D.20 Peripheral Link')
        print('\t', 'Report Deadband not in XML')
        return

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b013_check(app):
    # Check Reset Link on Rx NACK
    # Check DCD, RTS, & CTS
    # Check DCD to RX Enable Time
    # Check Baud Rate
    # Check RTS Preamble
    # Check RTS Postamble
    # Check Max Frame Size
    # Check Transmit Retries
    # Check Transmit Buffers
    # Check Receive Buffers
    # Check Confirm Timeout
    # Check Response Timeout

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'), ':', app[0].get('Table_Name'), 'Table')
        # Port Com Global Variable, enable editing
        set_comlist()
        # Loop through the Port Configuration table
        for i, record in enumerate(app[0]):
            print('\t\t', i, ':')
            print('\t\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # Port
            print('\t\t\t', record[2].get('Field_Name'), ':', record[2].get('Field_Value'))  # Reset Link on RX NACK
            print('\t\t\t', record[3].get('Field_Name'), ':', record[3].get('Field_Value'))  # DCD
            print('\t\t\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))  # RTS
            print('\t\t\t', record[5].get('Field_Name'), ':', record[5].get('Field_Value'))  # CTS
            print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))  # DCD to Rx Enable Time
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # Baud Rate
            print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))  # RTS Preamble
            print('\t\t\t', record[8].get('Field_Name'), ':', record[8].get('Field_Value'))  # RTS Postamble
            print('\t\t\t', record[9].get('Field_Name'), ':', record[9].get('Field_Value'))  # Max Frame Size
            print('\t\t\t', record[10].get('Field_Name'), ':', record[10].get('Field_Value'))  # Transmit Retries
            print('\t\t\t', record[11].get('Field_Name'), ':', record[11].get('Field_Value'))  # Transmit Buffers
            print('\t\t\t', record[12].get('Field_Name'), ':', record[12].get('Field_Value'))  # Receive Buffers
            print('\t\t\t', record[13].get('Field_Name'), ':', record[13].get('Field_Value'))  # Confirm Timeout
            print('\t\t\t', record[14].get('Field_Name'), ':', record[14].get('Field_Value'))  # Response Timeout

            # Append Port Com to list
            b013_com_list.append((record[0].get('Field_Value')))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b014_check(app):
    # Check SOE BUFFER SIZE = 500
    # Check SOE LOCATION = NVRAM
    # Check User Name = something
    # Check Password = something
    # Check Control Password = something

    # Check if Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[3].get('Table_Identifier'), ':', app[3].get('Table_Name'), 'Table')
        # Loop through the Buffer Configuration table
        for i, record in enumerate(app[3]):
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # SOE Buffer Size
            print('\t\t', record[4][0][0][0].get('Field_Name'), ':', record[4][0][0][0].get('Field_Value'))
                                                                                           # SOE Location

        # Print the table identifier followed by the table name for clarity
        print('\t', app[4].get('Table_Identifier'), ':', app[4].get('Table_Name'), 'Table')
        # Loop through the User Configuration table
        for i, record in enumerate(app[4]):
            record_num = int(record.get('Record_Number')) - 1
            print('\t\t', 'Record ', record_num, ':')
            print('\t\t\t', record[5].get('Field_Name'), ':', record[5].get('Field_Value'))  # User Name
            print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))  # Password
            print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))  # Control Password

        # Print the table identifier followed by the table name for clarity
        print('\t', app[7].get('Table_Identifier'), ':', app[7].get('Table_Name'), 'Table')
        # Loop through the Welcome Message table
        for i, record in enumerate(app[7]):
            print('\t\t', record[0].get('Field_Name'), record[0].get('Field_Value'), record[3].get('Field_Name'), ':',
                  record[3].get('Field_Value'))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b015_check(app):
    # Check Bridgeman app

    # Check if Application in Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Count the number of remote DNP devices
        num_dnp_dev = 0
        for record in app[5]:
            num_dnp_dev += 1
        print('\t', num_dnp_dev, 'remote DNP devices')
        print('\t', app[0][0][1].get('Field_Name'), ':', app[0][0][1].get('Field_Value'))  # Number of Rx Buffers

        print('\t', 'Local Application Table [LAN Address(Hex), Data Link channel]')
        # Loop through the Local Application table
        for i, record in enumerate(app[3]):
            print('\t\t', record[0].get('Field_Value'), '(x', record[3].get('Field_Value'), ')',
                  record[2].get('Field_Value'), ':', b013_com_list[i])

        print('\t', 'Remote Application Table [LAN Address(Hex), Data Link channel]')
        # Loop through the Remote Application table
        for record in app[5]:
            print('\t\t', record[0].get('Field_Value'), '(x', record[3].get('Field_Value'), ')',
                  record[2].get('Field_Value'), '   -   ', record[4].get('Field_Name'),
                  ':', record[4].get('Field_Value'))

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b021_check(app):
    # Check Datalink Confirm
    # Check Time Sync Enable State
    # Check Offline Local IIN
    # Check Idle Report Period
    # Check SOE

    # Check if the Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'), ':', app[0].get('Table_Name'), 'Table')
        # Loop through the DPA Configuration table
        for i, record in enumerate(app[0]):
            print('\t\t', i, ':')
            print('\t\t\t', record[15].get('Field_Name'), ':', record[15].get('Field_Value'))  # Data Link Confirm
            print('\t\t\t', record[39].get('Field_Name'), ':', record[39].get('Field_Value'))  # Time Sync Enable State
            print('\t\t\t', record[25][0][0][3].get('Field_Name'), ':', record[25][0][0][3].get('Field_Value'))
                                                                                               # Offline Sets Local IIN
            print('\t\t\t', record[12][0][0][5].get('Field_Name'), ':', record[12][0][0][5].get('Field_Value'))
                                                                                               # Idle Report Period

        # Print the table identifier followed by the table name for clarity
        print('\t', app[3].get('Table_Identifier'), ':', app[3].get('Table_Name'), 'Table')
        counter = 0
        for i, record in enumerate(app[3]):
            # Loop through the Binary Input Map table
            if record[3].get('Field_Value') == 'Enabled':
                pass
            else:
                counter = counter + 1  # For tracking purposes

        # If the values are all enabled, print statement
        if counter == 0:
            print('\t\t', app[3][0][3].get('Field_Name'), ':', app[3][0][3].get('Field_Value'))
        else:
            print('\t\t', 'An', app[3][0][3].get('Field_Name'), 'value is not enabled. Please check the SGConfig.')

        # Compare the Winpoints from the points list to what's programmed in the D20

        # Put in the path to the excel template file
        directory = os.path.expanduser(
            os.path.join('~', 'Documents', 'GitHub', 'FE-D20_Checker', 'Example D20 XML', 'D20MEII'))

        # Determine that the XCEL template's filename is D20 DNP Map WinPt Check
        for thing in os.listdir(directory):
            # Excel template should be named D20 DNP Map WinPt Check
            if 'D20 DNP Map WinPt Check' in thing:
                xcel_filename = thing
                print('\t', xcel_filename)

            # Try to read the file
            try:
                # Add the filename to the directory
                filepath = directory + '/' + xcel_filename

                # Open the excel document for reading
                wbook = xlrd.open_workbook(filepath)

                # Read the specified excel sheet
                for sheet in wbook.sheet_names():
                    if 'Sheet1' in sheet:
                        wsheet_name = sheet

                wsheet = wbook.sheet_by_name(wsheet_name)

                # Determine which column the specified points are in
                for i, cell in enumerate(wsheet.row(1)):
                    if cell.value == 'DNP INDEX':
                        dnp_index = i
                    elif cell.value == 'STATUS':
                        status_index = i
                    elif cell.value == 'ANALOG':
                        analog_index = i
                    elif cell.value == 'CONTROL':
                        control_index = i

            # PyCharm presents an error if the excel file is open. You have to close the document
            # before running the program
            except Exception:
                print('\t\t\t', 'Error: Cannot find the file.')

        # Call the WinPt check function for Status, Analog, and Control points respectively
        winpt_check(xcel_filename, directory, app, status_index, 3, 'Status')
        winpt_check(xcel_filename, directory, app, analog_index, 6, 'Analog')
        winpt_check(xcel_filename, directory, app, control_index, 4, 'Control')

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')

def b023_check(app):
    # <app>
    #   <table "B023_PNT">
    #   <table "B023_POL">
    #   <table "B023_DEV">
    #   <table "B023_CFG">

    # Check if Application is Enabled
    if app.get('Enabled') == 'True':
        # Print the application identifier followed by the application name for clarity
        print(app.get('Application_Identifier'), '-', app.get('Application_Name'))
        # Print the table identifier followed by the table name for clarity
        print('\t', app[2].get('Table_Identifier'))
        b023_pnt_list = []

        # Loop through the Device Point Map table
        for i, record in enumerate(app[2]):
            # Analog Input
            print('\t\t', i, '-', record[0].get('Field_Value'), ':', record[1].get('Field_Value'))
            b023_pnt_list.append((record[0].get('Field_Value'), record[1].get('Field_Value')))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[3].get('Table_Identifier'))
        b023_pol_list = []
        # Loop through the Device Poll table
        for i, record in enumerate(app[3]):
            print('\t\t', i, ':')
            print('\t\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))  # Poll Data Type
            print('\t\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # Qualifier
            if record[4].get('Field_Value') != 0:
                print('\t\t\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))  # Poll Interval (Days)
            if record[5].get('Field_Value') != 0:
                print('\t\t\t', record[5].get('Field_Name'), ':', record[5].get('Field_Value'))  # Poll Interval (Hours)
            if record[6].get('Field_Value') != 0:
                print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))  # Poll Interval (Minutes)
            if record[7].get('Field_Value') != 0:
                print('\t\t\t', record[7].get('Field_Name'), ':', record[7].get('Field_Value'))  # Poll Interval (Seconds)
            if record[8].get('Field_Value') != 0:
                print('\t\t\t', record[8].get('Field_Name'), ':', record[8].get('Field_Value'))  # Poll Interval (Msec)
            # Append to the b023 pol list
            b023_pol_list.append((i, record[0].get('Field_Value')))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[1].get('Table_Identifier'))
        b023_dev_list = []
        # Loop through the Device Configuration table
        for i, record in enumerate(app[1]):
            # Application Address
            print('\t\t', record[0].get('Field_Name'), ':', record[0].get('Field_Value'))
            # Data Link CFM Rquired
            print('\t\t\t', record[3][0][0][7].get('Field_Name'), ':', record[3][0][0][7].get('Field_Value'))
            # Off-Line After Fail
            print('\t\t\t', record[3][0][0][8].get('Field_Name'), ':', record[3][0][0][8].get('Field_Value'))
            # Time Sync Method
            print('\t\t\t', record[4].get('Field_Name'), ':', record[4].get('Field_Value'))
            # Failures For bad Channel
            print('\t\t\t', record[6].get('Field_Name'), ':', record[6].get('Field_Value'))
            # First Point Record
            print('\t\t\t', record[8][0][0][2].get('Field_Name'), ':', record[8][0][0][2].get('Field_Value'))
            # Number of Point Records
            print('\t\t\t', record[8][0][0][3].get('Field_Name'), ':', record[8][0][0][3].get('Field_Value'))
            # Analog Input
            for index in range(int(record[8][0][0][2].get('Field_Value')),
                               int(record[8][0][0][2].get('Field_Value')) + int(record[8][0][0][3].get('Field_Value'))):
                print('\t\t\t\t', b023_pnt_list[index])
            # First Poll Record
            print('\t\t\t', record[8][0][0][4].get('Field_Name'), ':', record[8][0][0][4].get('Field_Value'))
            # Number of Poll Records
            print('\t\t\t', record[8][0][0][5].get('Field_Name'), ':', record[8][0][0][5].get('Field_Value'))
            # Integrity Poll
            for index in range(int(record[8][0][0][4].get('Field_Value')),
                               int(record[8][0][0][4].get('Field_Value')) + int(record[8][0][0][5].get('Field_Value'))):
                print('\t\t\t\t', b023_pol_list[index])
            # Events for Time Sync
            print('\t\t\t', record[9][0][0][5].get('Field_Name'), ':', record[9][0][0][5].get('Field_Value'))
            # Append to the b023 dev list
            b023_dev_list.append(record[0].get('Field_Value'))

        # Print the table identifier followed by the table name for clarity
        print('\t', app[0].get('Table_Identifier'))
        # Loop through the DCA Configuration table
        for i, record in enumerate(app[0]):
            print('\t\t', record[1].get('Field_Name'), ':', record[1].get('Field_Value'))  # Application Address
            print('\t\t\t', record[3].get('Field_Name'), ':', record[3].get('Field_Value'))  # Min Inter Poll Delay
            print('\t\t\t', record[2][0][0][0].get('Field_Name'), ':', record[2][0][0][0].get('Field_Value'))
                                                                                           # Restart Delay
            print('\t\t\t', 'Devices in DCA:') # Devices in DCA
            for index in range(int(record[10].get('Field_Value')),
                               int(record[10].get('Field_Value')) + int(record[11].get('Field_Value'))):
                # Print the b023 dev list to indicate the number of devices in DCA
                print('\t\t\t\t', b023_dev_list[index])

    # If the application is disabled, print statement
    else:
        print(app.get('Application_Identifier'), '-', 'is disabled')
    return
